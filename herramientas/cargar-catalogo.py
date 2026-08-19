#!/usr/bin/env python3
"""
CONVERTIR TU PLANILLA EN EL CATÁLOGO DE LA WEB
==============================================

Sirve para cargar los productos desde Excel/Google Sheets en vez de editar
data/productos.js a mano. Pensado para catálogos grandes (+200 productos).

CÓMO USARLO
-----------
1. Abrí  data/productos-plantilla.csv  con Excel o Google Sheets.
2. Cargá tus productos (una fila por producto). Ver formato abajo.
3. Guardá / exportá como CSV con el nombre  data/productos.csv
4. Desde la carpeta de la tienda, ejecutá en la terminal:

       python3 herramientas/csv-a-productos.py

   Se regenera  data/productos.js  con tu catálogo real.
   (Se guarda una copia de seguridad del archivo anterior.)

FORMATO DE LAS COLUMNAS
-----------------------
id               REM-001            Único. No repetir.
codigo           4216               Código de artículo que ve el cliente.
nombre           Remera Oversize    Como se muestra en la web.
slug             remera-oversize    Sin espacios ni acentos. Vacío = se genera solo.
categoria        remeras            Tiene que existir en data/config.js
subcategoria     oversize           Tiene que existir dentro de esa categoría.
genero           mujer              "mujer", "hombre" o "unisex".
                                    Las unisex aparecen en las dos secciones.
precio           34900              Precio MINORISTA. Solo números.
precioMayorista  20900              Precio POR MAYOR. Solo números.
precioAnterior   45900              Precio tachado. Vacío si no está en oferta.
colores          negro|blanco       Separados por | . Deben existir en config.js
stock_XS ...     5                  Unidades por talle. 0 = agotado. Vacío = no aplica.
stock_Único      12                 Para accesorios y productos sin talle.
imagenes         img/productos/a.jpg|img/productos/b.jpg     Separadas por |
destacado        true / false
nuevo            true / false
descripcion      texto libre
materiales       texto libre
"""
import csv, json, os, re, sys, shutil, unicodedata

AQUI = os.path.dirname(os.path.abspath(__file__))
RAIZ = os.path.dirname(AQUI)
CSV_ENTRADA  = os.path.join(RAIZ, "data", "productos.csv")
XLSX_ENTRADA = os.path.join(RAIZ, "data", "productos-plantilla.xlsx")
JS_SALIDA    = os.path.join(RAIZ, "data", "productos.js")


def filas_desde_excel(ruta):
    """Lee la planilla de Excel. Los encabezados están en la fila 3."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("✗ Falta la librería openpyxl. Instalala con:")
        print("     pip install openpyxl")
        sys.exit(1)

    wb = load_workbook(ruta, data_only=True)
    ws = wb["Productos"] if "Productos" in wb.sheetnames else wb.worksheets[0]

    encabezados = [str(c.value).strip() if c.value else "" for c in ws[3]]
    filas = []
    for fila in ws.iter_rows(min_row=4, values_only=True):
        if not any(v not in (None, "") for v in fila):
            continue
        d = {}
        for k, v in zip(encabezados, fila):
            if k:
                d[k] = "" if v is None else str(v)
        # La fila 4 de la plantilla es el ejemplo: se saltea
        if d.get("codigo", "").strip() == "4216" and "Oversize Algod" in d.get("nombre", ""):
            continue
        filas.append(d)
    return filas


def filas_desde_csv(ruta):
    with open(ruta, newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))
TALLES = ["XS", "S", "M", "L", "XL", "XXL", "Único"]

CABECERA = """/* ==========================================================================
   CATÁLOGO DE PRODUCTOS
   Generado automáticamente desde data/productos.csv
   (no edites este archivo a mano si trabajás con la planilla)
   ========================================================================== */

window.PRODUCTOS = """


def slugify(t):
    t = unicodedata.normalize("NFKD", t).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z0-9]+", "-", t.lower()).strip("-")


def a_bool(v):
    return str(v).strip().lower() in ("1", "true", "si", "sí", "x", "verdadero")


def a_int(v, por_defecto=0):
    v = re.sub(r"[^\d-]", "", str(v or ""))
    return int(v) if v else por_defecto


def main():
    # Se acepta el Excel (preferido) o un CSV
    if os.path.exists(XLSX_ENTRADA):
        origen = XLSX_ENTRADA
        filas_crudas = filas_desde_excel(XLSX_ENTRADA)
    elif os.path.exists(CSV_ENTRADA):
        origen = CSV_ENTRADA
        filas_crudas = filas_desde_csv(CSV_ENTRADA)
    else:
        print("✗ No encontré la planilla. Tiene que estar una de estas dos:")
        print(f"     {XLSX_ENTRADA}")
        print(f"     {CSV_ENTRADA}")
        sys.exit(1)

    print(f"• Leyendo: {os.path.basename(origen)}  ({len(filas_crudas)} filas con datos)")
    productos, ids, errores, avisos = [], set(), [], []

    if True:
        for n, fila in enumerate(filas_crudas, start=5):
            nombre = (fila.get("nombre") or "").strip()
            codigo = (fila.get("codigo") or "").strip()
            # Si no hay columna "id", se usa el código como identificador
            pid = (fila.get("id") or codigo).strip()
            if not pid and not nombre:
                continue
            if not pid:
                errores.append(f"fila {n}: falta el código"); continue
            if pid in ids:
                errores.append(f"fila {n}: id repetido '{pid}'"); continue
            ids.add(pid)

            stock = {}
            for t in TALLES:
                col = f"stock_{t}"
                if col in fila and str(fila[col]).strip() != "":
                    stock[t] = a_int(fila[col])
            if not stock:
                stock = {t: 0 for t in TALLES[:6]}

            precio = a_int(fila.get("precio"))
            if precio <= 0:
                errores.append(f"fila {n}: precio inválido en '{pid}'")

            mayorista = a_int(fila.get("precioMayorista"), 0)
            if mayorista <= 0:
                mayorista = int(precio * 0.65)
                avisos.append(f"'{pid}': sin precio mayorista, se calculó el 65% ({mayorista})")
            elif mayorista >= precio:
                avisos.append(f"'{pid}': el precio mayorista NO es menor al minorista")

            anterior = a_int(fila.get("precioAnterior"), 0) or None

            genero = (fila.get("genero") or "").strip().lower()
            if genero in ("m", "mujer", "dama", "femenino"):
                genero = "mujer"
            elif genero in ("h", "hombre", "caballero", "masculino"):
                genero = "hombre"
            elif genero in ("u", "unisex", ""):
                if not genero:
                    avisos.append(f"'{pid}': sin género, se cargó como unisex")
                genero = "unisex"
            else:
                avisos.append(f"'{pid}': género '{genero}' no reconocido, se cargó como unisex")
                genero = "unisex"

            productos.append({
                "id": pid,
                "codigo": (fila.get("codigo") or pid).strip(),
                "nombre": nombre,
                "slug": (fila.get("slug") or "").strip() or slugify(nombre) + "-" + slugify(pid),
                "categoria": (fila.get("categoria") or "").strip().lower(),
                "subcategoria": (fila.get("subcategoria") or "").strip().lower(),
                "genero": genero,
                "precio": precio,
                "precioMayorista": mayorista,
                "precioAnterior": anterior,
                "colores": [c.strip().lower() for c in (fila.get("colores") or "").split("|") if c.strip()],
                "stock": stock,
                # Si solo pusieron el nombre del archivo, se le agrega la carpeta
                "imagenes": [(i.strip() if "/" in i else "img/productos/" + i.strip())
                             for i in (fila.get("imagenes") or "").split("|") if i.strip()],
                "destacado": a_bool(fila.get("destacado")),
                "nuevo": a_bool(fila.get("nuevo")),
                "descripcion": (fila.get("descripcion") or "").strip(),
                "materiales": (fila.get("materiales") or "").strip(),
                "origen": (fila.get("origen") or "Confeccionado en Argentina").strip(),
                "fechaAlta": (fila.get("fechaAlta") or "2026-01-01").strip(),
            })

    if errores:
        print("✗ Errores (esas filas NO se cargaron):")
        for e in errores:
            print("   -", e)
        print()
    if avisos:
        print("⚠  Avisos:")
        for a in avisos[:20]:
            print("   -", a)
        if len(avisos) > 20:
            print(f"   ... y {len(avisos) - 20} más")
        print()

    if not productos:
        print("✗ No se cargó ningún producto. Revisá el CSV.")
        sys.exit(1)

    if os.path.exists(JS_SALIDA):
        shutil.copyfile(JS_SALIDA, JS_SALIDA + ".backup")
        print("• Copia de seguridad: data/productos.js.backup")

    with open(JS_SALIDA, "w", encoding="utf-8") as f:
        f.write(CABECERA + json.dumps(productos, ensure_ascii=False, indent=2) + ";\n")

    sin_foto = sum(1 for p in productos if not p["imagenes"])
    print(f"✓ Listo: {len(productos)} productos escritos en data/productos.js")
    if sin_foto:
        print(f"  ({sin_foto} todavía sin fotos cargadas)")


if __name__ == "__main__":
    main()
